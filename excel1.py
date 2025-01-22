from openpyxl import load_workbook
workbook=load_workbook("D:/Excel/writeme.xlsx")
sheet=workbook["Sample Excel Sheet"]
#sheet["C1"] = "new data in C1"
sheet.cell(row=1,column=3,value="CITY")
workbook.save("D:/Excel/writeme.xlsx")

